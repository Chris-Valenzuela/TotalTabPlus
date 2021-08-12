
from scraper import *
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def makeup(totaltabsdf, newcolumns):

    #======================  Styling and openpyxl ===============================
    book = load_workbook(filename)
    finalsheetname = "TotalTabPlus"
    book.create_sheet(finalsheetname,index = 2)

    writer = pd.ExcelWriter(output, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


    totaltabsdf.to_excel(writer, finalsheetname, startrow = 0)
    ws = writer.sheets['TotalTabPlus']
    # this code below adds a filter to the top row - print(ws.dimensions) print Topleft:BottomRight
    ws.auto_filter.ref = ws.dimensions


    openpyxlcollet = []

    for col in range(0, len(newcolumns)):
        if newcolumns[col].find("Max Diff") != -1:
            #print(get_column_letter(col + 2))
            openpyxlcollet.append(get_column_letter(col + 1))
            

    endrow = len(totaltabsdf) + 1
    for letter in openpyxlcollet:
        ws.conditional_formatting.add(letter+'2:'+ letter + str(endrow),
                        ColorScaleRule(start_type='percentile', start_value=0, start_color='AA0000',
                                    mid_type='percentile', mid_value=50, mid_color='f7f700',
                                    end_type='percentile', end_value=100, end_color='00aa00')
                                    )
        
    for x, col in enumerate(totaltabsdf):    
        #Font colorized all the cells with data and not max diff. 
        if col != 'Table' and col != 'Question' and col != 'Stub' and col != 'TableLink' and col.find('Max Diff') == -1:
            for y, rowdata in enumerate(totaltabsdf[col]):
                ws[get_column_letter(x + 2) + str(y + 2)].font = Font( color = bough.color_font(str(rowdata))) 




    # ============== Creating hyper links for the "Table Link" Column ========================

    rownumber = len(newcolumns)

    # grab the Table Name and makes it a hyper link
    ws = book['TotalTabPlus']
    for x, i in enumerate(range(1, ws.max_row + 1 )):
        x += 1    
        linkStr =  ws.cell(row=i, column = rownumber).value
        if len(linkStr.split(" ")) > 1:
            TableNumber = linkStr.split(" ")[1]         
            link = '#T' + str(TableNumber) + "!" + 'A1'
            ws.cell(row=x, column = rownumber).value =  '=HYPERLINK("{}", "{}")'.format(link, ws.cell(row=i, column = rownumber).value)
            ws.cell(row=x, column = rownumber).style = "Hyperlink"
            

    writer.save()  